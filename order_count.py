#!/usr/bin/env python3
"""
order_count.py —— 从品项销售统计表里，按"辣度/口味选择 = 一单"的规则
反推出 堂食 / 外卖 的订单数和日均单量。

用法：
    python3 order_count.py                        # 默认读 data/sales_raw.xlsx
    python3 order_count.py --file data/xxx.xlsx   # 指定文件
    python3 order_count.py --days 40              # 手动指定统计天数
    python3 order_count.py --csv out.csv          # 同时导出明细 CSV

背景与规则：
- 每单必选一个"辣度/口味"，所以"辣度行的销量"就等于"订单数"。
- 不同渠道的辣度写法不一样：
    * 店内点餐/团购/微信：  "不辣" / "微辣" / "中辣" / "特辣" / "微微辣"
    * 美团外卖/淘宝闪购：   "【口味】不辣" …
    * 京东秒送：            "不辣（含包装打包费）" …
- 订单子来源 → 堂食/外卖 分类见 CATEGORY。
"""

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl  # noqa
except ImportError:
    sys.exit("请先安装 openpyxl： pip install openpyxl")

# ---- 配置 ----
TASTES = ['微微辣', '不辣', '中辣', '特辣', '微辣']  # 微微辣优先，避免被"微辣"子串吞掉

# 匹配三种写法：纯辣度 / 【口味】辣度 / 辣度（含包装打包费）
TASTE_PAT = re.compile(
    r'^(?:【口味】)?(微微辣|不辣|中辣|特辣|微辣)(?:（含包装打包费）)?$'
)

# 订单子来源 → 业务类别
CATEGORY = {
    '店内点餐':     '堂食-店内',
    '微信小程序':   '堂食-店内',        # 量极少，并入堂食
    '美团点评团购': '堂食-团购核销',
    '抖音团购':     '堂食-团购核销',
    '美团外卖':     '外卖',
    '淘宝闪购':     '外卖',
    '京东秒送':     '外卖',
}


def parse_args():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--file', default='data/sales_raw.xlsx',
                    help='品项销售统计 xlsx 路径（默认 data/sales_raw.xlsx）')
    ap.add_argument('--days', type=int, default=None,
                    help='统计天数；不填则尝试从表头"营业日期【start-end】"自动解析')
    ap.add_argument('--csv', default=None, help='可选：输出明细 CSV')
    return ap.parse_args()


def auto_days(ws):
    """从第 2 行那段长文本里解析 营业日期【2026/04/01-2026/05/10】 的天数。"""
    txt = ws.cell(row=2, column=1).value or ''
    m = re.search(r'营业日期【(\d{4}/\d{1,2}/\d{1,2})-(\d{4}/\d{1,2}/\d{1,2})】', txt)
    if not m:
        return None
    from datetime import datetime
    fmt = '%Y/%m/%d'
    d1 = datetime.strptime(m.group(1), fmt)
    d2 = datetime.strptime(m.group(2), fmt)
    return (d2 - d1).days + 1


def main():
    args = parse_args()
    path = Path(args.file)
    if not path.exists():
        sys.exit(f'文件不存在：{path}')

    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[3]]
    try:
        idx_src  = header.index('订单子来源')
        idx_name = header.index('菜品名称')
        idx_qty  = header.index('销售数量')
    except ValueError as e:
        sys.exit(f'表头缺列：{e}（期望：订单子来源/菜品名称/销售数量）')

    days = args.days or auto_days(ws) or 1
    if days <= 0:
        sys.exit('统计天数必须 > 0')

    # src -> taste -> qty
    cross = defaultdict(lambda: defaultdict(float))
    for row in ws.iter_rows(min_row=5, values_only=True):
        name = row[idx_name]; src = row[idx_src]; qty = row[idx_qty]
        if not name or qty is None or src is None:
            continue
        m = TASTE_PAT.match(str(name).strip())
        if m:
            cross[src][m.group(1)] += qty

    # ---- 输出 ----
    print(f'文件：{path}    统计天数：{days} 天\n')

    # 明细表
    print('=== 订单子来源 × 辣度  单数明细 ===')
    head = f'{"订单子来源":<14}{"类别":<18}' + ''.join(f'{t:>8}' for t in TASTES) + f'{"合计":>10}{"日均":>10}'
    print(head); print('-' * len(head))
    srcs = sorted(cross.keys(), key=lambda s: -sum(cross[s].values()))
    detail_rows = []
    for src in srcs:
        d = cross[src]
        vals = [d.get(t, 0) for t in TASTES]
        total = sum(vals)
        cat = CATEGORY.get(src, '?')
        print(f'{src:<14}{cat:<18}' + ''.join(f'{v:>8.0f}' for v in vals)
              + f'{total:>10.0f}{total/days:>10.1f}')
        detail_rows.append((src, cat, *vals, total, total/days))

    # 汇总
    bucket = defaultdict(float)
    for src, d in cross.items():
        cat = CATEGORY.get(src, '未分类')
        bucket[cat] += sum(d.values())
    dine_in = sum(v for k, v in bucket.items() if k.startswith('堂食'))
    delivery = bucket.get('外卖', 0)

    print(f'\n=== 堂食 / 外卖 汇总（{days} 天）===')
    print(f'  🏠 堂食（店内+团购核销）：{dine_in:>6.0f} 单 → 日均 {dine_in/days:>5.1f} 单/天')
    print(f'  🛵 外卖（美团+淘宝闪购+京东秒送）：{delivery:>6.0f} 单 → 日均 {delivery/days:>5.1f} 单/天')
    total_all = dine_in + delivery
    if total_all:
        print(f'  📊 合计：{total_all:.0f} 单 → 日均 {total_all/days:.1f} 单/天'
              f'（堂食占比 {dine_in/total_all:.1%}，外卖 {delivery/total_all:.1%}）')

    # 外卖内部拆分
    print('\n=== 外卖平台明细 ===')
    for src in ['美团外卖', '淘宝闪购', '京东秒送']:
        t = sum(cross.get(src, {}).values())
        print(f'  {src:<10}  {t:>5.0f} 单 → 日均 {t/days:>4.1f} 单/天')

    if args.csv:
        import csv
        with open(args.csv, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(['订单子来源', '类别', *TASTES, '合计单数', '日均单数'])
            w.writerows(detail_rows)
        print(f'\n✅ 明细已写入 {args.csv}')


if __name__ == '__main__':
    main()
