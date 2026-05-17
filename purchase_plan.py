#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据盘货表 + SKU 字典 + 销售摘要，计算补货建议。

【核心规则（老王 2026-05-12 更新：每周两次到货）】
  到货日：周一、周四。两次到货之间的空档：
    - 周一→周四 = 3 天
    - 周四→下周一 = 4 天（含周末高峰）

  下单量目标 = 下次到货前的预计消耗 + 安全 buffer（默认 0.5 天）
  目标库存   = 日均消耗 × (gap_days + buffer_days)
  缺口       = max(0, 目标 - 当前)
  下单箱数   = ceil(缺口 / 每箱基准量)

输入：
  data/sku_dictionary.csv    —— SKU 主数据（含进货规格、袋大小）
  data/sku_sales_summary.csv —— 销售摘要（含日均消耗）
  /tmp/stock_pan.xlsx        —— 当前盘货（按袋）

输出：
  data/purchase_plan.csv     —— 详细报表
  控制台汇总

用法:
  python3 purchase_plan.py                      # 自动按今天周几判断下次到货
  python3 purchase_plan.py --gap 3              # 手动指定到下次到货间隔天数
  python3 purchase_plan.py --buffer 1           # 安全 buffer 天数（默认 0.5）
  python3 purchase_plan.py --weekend-boost 1.2  # 周末销量加成（跨周末时生效）
  python3 purchase_plan.py --stock /path/to/stock.xlsx
  python3 purchase_plan.py --days 5             # 【兼容旧用法】直接指定总天数
"""
import argparse
import csv
import datetime as dt
import json
import math
import sys
from pathlib import Path

DATA = Path(__file__).parent / "data"

# 到货日：0=周一, 3=周四
DELIVERY_WEEKDAYS = [0, 3]


def next_delivery_gap(today: dt.date) -> tuple[int, bool]:
    """
    计算从今天到"下一次到货日"的天数；以及本次覆盖期是否跨周末。
    约定：今天即使是到货日，也算"今天订今天到"，看到下一次到货。
    """
    wd = today.weekday()  # 0=Mon ... 6=Sun
    # 找下一个 delivery weekday（严格 > 今天）
    candidates = [d for d in DELIVERY_WEEKDAYS if d > wd]
    if candidates:
        gap = min(candidates) - wd
    else:
        # 今天之后本周没有到货日了 → 下周一
        gap = 7 - wd + DELIVERY_WEEKDAYS[0]

    # 覆盖期内是否包含周六或周日
    cross_weekend = False
    for i in range(1, gap + 1):
        if (wd + i) % 7 in (5, 6):   # 5=Sat, 6=Sun
            cross_weekend = True
            break
    return gap, cross_weekend


# ============ CLI ============
ap = argparse.ArgumentParser()
ap.add_argument("--gap", type=float, default=None,
                help="到下次到货的天数（不给就按今天自动算）")
ap.add_argument("--buffer", type=float, default=0.5,
                help="安全 buffer 天数（默认 0.5 天）")
ap.add_argument("--weekend-boost", type=float, default=1.15,
                help="覆盖期跨周末时的销量加成系数（默认 1.15）")
ap.add_argument("--seasoning-days", type=float, default=10,
                help="调料类特殊覆盖天数（默认 10 天，因为频次低且购买粒度大）")
ap.add_argument("--days", type=float, default=None,
                help="【兼容】直接指定总目标天数，给了就覆盖 gap+buffer")
ap.add_argument("--stock", default="/tmp/stock_pan.xlsx",
                help="盘货 xlsx 路径")
ap.add_argument("--out", default=str(DATA / "purchase_plan.csv"),
                help="输出 CSV 路径")
ap.add_argument("--today", default=None,
                help="模拟日期 YYYY-MM-DD（测试用）")
args = ap.parse_args()

# 推断目标天数
today = dt.date.fromisoformat(args.today) if args.today else dt.date.today()
WEEKDAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

if args.days is not None:
    TARGET_DAYS = args.days
    GAP = args.days
    BUFFER = 0
    CROSS_WEEKEND = False
    MODE = f"手动指定总天数={args.days}"
else:
    if args.gap is not None:
        GAP = args.gap
        # 粗略判断：gap>=3 且跨周四/周日的默认当跨周末
        CROSS_WEEKEND = GAP >= 4
    else:
        GAP, CROSS_WEEKEND = next_delivery_gap(today)
    BUFFER = args.buffer
    TARGET_DAYS = GAP + BUFFER
    MODE = (
        f"今天 {today} ({WEEKDAY_NAMES[today.weekday()]}) → "
        f"下次到货还有 {GAP:g} 天，buffer {BUFFER:g} 天"
        + ("（跨周末）" if CROSS_WEEKEND else "")
    )

SEASONING_DAYS = args.seasoning_days

WEEKEND_MULT = args.weekend_boost if CROSS_WEEKEND else 1.0
STOCK_XLSX = args.stock


# ============ 加载 SKU 字典 ============
sku_map = {}
with (DATA / "sku_dictionary.csv").open(encoding="utf-8-sig") as f:
    for r in csv.DictReader(f):
        sku_map[r["sku_id"]] = r


# ============ 加载销售摘要（日均消耗） ============
# 注意：sku_sales_summary.csv 顶部有 # 开头的元数据行，要跳过
daily_use = {}
for path in [DATA / "sku_sales_summary.csv"]:
    if not path.exists():
        continue
    with path.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(line for line in f if not line.startswith("#"))
        for r in reader:
            try:
                daily_use[r["sku_id"]] = float(r["日均消耗"])
            except (KeyError, ValueError, TypeError):
                pass

# ============ 加载调料日均消耗（seasoning_calc.py 生成） ============
seasoning_path = DATA / "seasoning_daily.csv"
if seasoning_path.exists():
    with seasoning_path.open(encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            v = (r.get("日均消耗_g") or "").strip()
            if not v:
                continue
            try:
                daily_use[r["sku_id"]] = float(v)
            except (ValueError, TypeError):
                pass

# ============ 加载包材日均消耗（手填） ============
packaging_path = DATA / "packaging_daily.csv"
if packaging_path.exists():
    with packaging_path.open(encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            v = (r.get("日均消耗") or "").strip()
            if not v:
                continue
            try:
                daily_use[r["sku_id"]] = float(v)
            except (ValueError, TypeError):
                pass


def box_size_of(row):
    """返回 1 箱 = 多少基准单位（通过 purchase_units 逐层换算）"""
    try:
        g = json.loads(row["purchase_units"])
    except Exception:
        return None

    def resolve(unit, seen=None):
        if seen is None:
            seen = set()
        if unit == row["base_unit"]:
            return 1.0
        if unit in seen:
            return None
        seen.add(unit)
        if unit not in g:
            return None
        edge = g[unit]
        sub = resolve(edge["to"], seen)
        if sub is None:
            return None
        return float(edge["rate"]) * sub

    return resolve("箱")


def bags_per_box(row):
    """返回 1 箱 = 多少袋"""
    box = box_size_of(row)
    pack = row.get("pack_size_base")
    try:
        pack = float(pack)
    except Exception:
        return None
    if not box or not pack:
        return None
    return int(round(box / pack))


# ============ 加载盘货 ============
import openpyxl

wb = openpyxl.load_workbook(STOCK_XLSX)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

# 自动定位表头：第一行以 'sku_id' 开头的才是真正表头
# （容错：导出的 xlsx 可能顶部多一行"文档标题"）
header_idx = None
for i, r in enumerate(rows):
    if r and str(r[0]).strip() == "sku_id":
        header_idx = i
        break
if header_idx is None:
    sys.exit(f"❌ 盘货表 {STOCK_XLSX} 找不到以 'sku_id' 开头的表头行")
header = rows[header_idx]
stock_data = []
for r in rows[header_idx + 1:]:
    if r and r[0]:  # 跳过空行
        stock_data.append(dict(zip(header, r)))


# ============ 计算补货 ============
plan = []
unknown_pack = []
no_sales = []

for s in stock_data:
    sku_id = s["sku_id"]
    if sku_id not in sku_map:
        continue
    sku = sku_map[sku_id]
    # 兼容新旧列名：旧 "当前库存_袋（请填）" / 新 "当前库存_件（请填）"
    current_bags = (s.get("当前库存_件（请填）")
                    or s.get("当前库存_袋（请填）"))

    pack = sku.get("pack_size_base", "")
    # 如果 pack_size_base 空（多见于包材：中间单位是 捆/提 而不是 袋），
    # 从 purchase_units JSON 里取"袋"的同位件（捆/提/条/袋）作为"件大小"
    if pack in ("", None):
        try:
            g = json.loads(sku["purchase_units"])
            base = sku["base_unit"]
            # 在 JSON 里找从"箱"下一级到 base_unit 的那一层
            mid = g.get("箱", {}).get("to")
            if mid and mid in g and g[mid].get("to") == base:
                pack = float(g[mid]["rate"])
        except Exception:
            pass
    if pack in ("", None) or current_bags in (None, ""):
        unknown_pack.append({
            "sku_id": sku_id, "sku_name": sku["sku_name"],
            "category": sku["category"], "reason": "进货规格未知或未盘货",
        })
        continue

    pack = float(pack)
    current_base = float(current_bags) * pack

    # —— 新规则：目标库存 = 日均 × (gap + buffer) × 周末加成 ——
    # 例外：
    #   调料   → SEASONING_DAYS（默认 10 天，不加周末倍率）
    #   包材   → 和食材一起进货，TARGET_DAYS，但不加周末加成（不跟销量波动）
    d = daily_use.get(sku_id)
    cat = sku.get("category")
    is_seasoning = cat == "调料"
    is_packaging = cat == "包材"
    if d is None or d <= 0:
        target_base = float(sku.get("safe_stock_base") or 0)
        target_source = "safe_stock(无销售数据)"
        effective_days = None
        no_sales.append(sku_id)
    elif is_seasoning:
        # 调料：不跨周末加成，用独立的 SEASONING_DAYS
        effective_days = SEASONING_DAYS
        target_base = d * effective_days
        target_source = f"{d:g}g/天×{SEASONING_DAYS:g}天(调料)"
    elif is_packaging:
        # 包材：和食材同口径 TARGET_DAYS，不加周末倍率
        effective_days = TARGET_DAYS
        target_base = d * effective_days
        target_source = f"{d:g}/天×{TARGET_DAYS:g}天(包材)"
    else:
        effective_days = TARGET_DAYS * WEEKEND_MULT
        target_base = d * effective_days
        if WEEKEND_MULT != 1.0:
            target_source = f"{d:g}/天×{TARGET_DAYS:g}天×{WEEKEND_MULT:g}(周末)"
        else:
            target_source = f"{d:g}/天×{TARGET_DAYS:g}天"

    deficit_base = max(0.0, target_base - current_base)
    days_left = (current_base / d) if (d and d > 0) else None

    box = box_size_of(sku)
    bpb = bags_per_box(sku)

    if deficit_base <= 0:
        action = "✅ 充足"
        boxes = 0
        note = f"可撑 {days_left:.1f} 天" if days_left is not None else ""
    else:
        if box and box > 0:
            boxes = math.ceil(deficit_base / box)
            replenish_base = boxes * box
            after_base = current_base + replenish_base
            after_days = (after_base / d) if (d and d > 0) else None
            bag_str = f"={boxes * bpb}袋" if bpb else ""
            note = (
                f"缺{deficit_base:.0f}{sku['base_unit']}, "
                f"进{boxes}箱={replenish_base:.0f}{sku['base_unit']}{bag_str}"
            )
            if after_days is not None:
                note += f", 进完可撑 {after_days:.1f} 天"
            action = "🛒 下单"
        else:
            boxes = 0
            action = "⚠️ 无箱规格"
            note = f"缺{deficit_base:.0f}{sku['base_unit']}, 但未配箱换算"

    plan.append({
        "sku_id": sku_id,
        "sku_name": sku["sku_name"],
        "category": sku["category"],
        "base_unit": sku["base_unit"],
        "日均消耗": f"{d:g}" if d else "",
        "目标天数": f"{TARGET_DAYS:g}" + (f"×{WEEKEND_MULT:g}" if WEEKEND_MULT != 1.0 else ""),
        "目标库存_基准": round(target_base, 1),
        "目标来源": target_source,
        "袋大小": pack,
        "当前_袋": current_bags,
        "当前_基准": current_base,
        "剩余天数": f"{days_left:.1f}" if days_left is not None else "",
        "缺口_基准": round(deficit_base, 1),
        "箱大小": box,
        "每箱袋数": bpb,
        "建议下单_箱": boxes,
        "建议下单_袋": boxes * bpb if bpb else "",
        "状态": action,
        "备注": note,
    })


# ============ 输出 CSV ============
out = Path(args.out)
with out.open("w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(plan[0].keys()))
    w.writeheader()
    w.writerows(plan)


# ============ 控制台打印 ============
print("=" * 110)
print(f"📦 补货建议：{MODE}")
if WEEKEND_MULT != 1.0:
    print(f"   ⚠️ 覆盖期跨周末，销量加成系数 × {WEEKEND_MULT:g}")
print(f"   油炸/蔬菜目标 = 日均 × {TARGET_DAYS:g}天"
      + (f" × {WEEKEND_MULT:g}" if WEEKEND_MULT != 1.0 else ""))
print(f"   调料目标   = 日均 × {SEASONING_DAYS:g}天（频次低，粒度大）")
print("=" * 110)

need_order = [p for p in plan if p["建议下单_箱"] > 0]
ok = [p for p in plan if p["状态"] == "✅ 充足"]
warn = [p for p in plan if p["状态"] == "⚠️ 无箱规格"]


def days_key(p):
    try:
        return float(p["剩余天数"])
    except Exception:
        return 999
need_order.sort(key=days_key)

print(f"\n🛒 需要下单 ({len(need_order)} 个 SKU):")
print(f"  {'SKU':<7} {'名称':<12} {'当前':>7} {'日均':>9} {'剩余天':>6} {'目标':>9}  {'下单':<20}")
print("  " + "-" * 92)
for p in need_order:
    cur_str = f"{int(p['当前_袋'])}袋"
    d_str = f"{p['日均消耗']}{p['base_unit']}" if p['日均消耗'] else "-"
    left = p['剩余天数'] or "-"
    target = f"{p['目标库存_基准']:g}{p['base_unit']}"
    order_str = f"{p['建议下单_箱']}箱={p['建议下单_袋']}袋"
    print(f"  {p['sku_id']:<7} {p['sku_name']:<12} {cur_str:>7} {d_str:>9} {left:>6} {target:>9}  {order_str:<20}")

print(f"\n✅ 库存充足 ({len(ok)} 个):")
for p in ok:
    cur_str = f"{int(p['当前_袋'])}袋"
    left = p['剩余天数'] or "-"
    print(f"  {p['sku_id']:<7} {p['sku_name']:<12} {cur_str:>7}  剩余 {left} 天")

if warn:
    print(f"\n⚠️ 未配箱换算 ({len(warn)} 个):")
    for p in warn:
        print(f"  {p['sku_id']:<7} {p['sku_name']:<12} {p['备注']}")

if unknown_pack:
    print(f"\n❓ 盘货未填/进货规格TBD ({len(unknown_pack)} 个):")
    for p in unknown_pack:
        print(f"  {p['sku_id']:<7} {p['sku_name']:<12} {p['category']:<6}  {p['reason']}")

if no_sales:
    print(f"\nℹ️ 无销售数据，已退回 safe_stock_base 作目标 ({len(no_sales)} 个): {', '.join(no_sales)}")

# 总结
print("\n" + "=" * 60)
print("📊 汇总")
print("=" * 60)
print(f"  今天日期:     {today} ({WEEKDAY_NAMES[today.weekday()]})")
print(f"  下次到货前:   {GAP:g} 天")
print(f"  安全 buffer:  {BUFFER:g} 天")
if WEEKEND_MULT != 1.0:
    print(f"  周末加成:     × {WEEKEND_MULT:g}")
print(f"  需下单 SKU:   {len(need_order)} 个")
print(f"  需下单总数:   {sum(p['建议下单_箱'] for p in need_order)} 箱")
print(f"  库存充足:     {len(ok)} 个")
print(f"  未盘/未配:    {len(warn) + len(unknown_pack)} 个")
print(f"\n📄 详细报表: {out}")
