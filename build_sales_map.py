#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 sales_raw.xlsx 构建 sales_map 草稿（v2）。
------------------------------------------------
输出:
  sales_map_draft.csv   —— 最终销售 → SKU 消耗映射（展开套餐/加工品后）
  ignored.csv           —— 忽略的行（辣度/饮料/盲盒/蔬菜缺规格等）
  need_decision.csv     —— 仍需老王决策的行（应该为 0）
  unmapped.csv          —— 脚本未能识别的行（应该为 0）
"""

from __future__ import annotations
import csv
import json
import re
import openpyxl
from collections import defaultdict
from pathlib import Path

DATA = Path("data")
SRC = DATA / "sales_raw.xlsx"
SKU_CSV = DATA / "sku_dictionary.csv"


def load_skus():
    skus = {}
    with SKU_CSV.open("r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            skus[row["sku_id"]] = row
    return skus


SKUS = load_skus()


def load_sales():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["品项销售统计"]
    rows = []
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 4).value
        if not name or name == "合计":
            continue
        rows.append({
            "source": ws.cell(r, 1).value,
            "category": ws.cell(r, 2).value,
            "item_type": ws.cell(r, 3).value,
            "name": name,
            "spec": ws.cell(r, 5).value or "",
            "unit": ws.cell(r, 6).value or "",
            "qty": ws.cell(r, 8).value or 0,
        })
    return rows


# ============================================================
#  映射规则
# ============================================================

# 单 SKU 映射：精确/包含匹配都在这
# 每条规则形如 (关键词, [(sku_id, base_qty 或 None), ...])
# base_qty=None 表示按规格/菜名自动解析；若指定则固定消耗量
NAME_TO_SKU = {
    # 串类（油炸品）—— 标准 1 串对 1 串
    "豆干小串": "SKU001", "卤香小豆干": "SKU001",
    "鱼豆腐串": "SKU002", "弹弹鱼豆腐": "SKU002",
    "掌中宝串": "SKU003", "掌中宝小串": "SKU003",
    "卤鸭肠小串": "SKU004", "爽脆鸭肠串": "SKU004", "劲脆鸭肠串": "SKU004",
    "玉米粒小串": "SKU005", "甜甜玉米粒": "SKU005",
    "五花肉串": "SKU006", "焦脆五花肉": "SKU006", "焦香五花肉": "SKU006",
    "纸豆皮小串": "SKU007", "薄脆纸豆皮": "SKU007", "薄脆豆皮串": "SKU007",
    "腊肠小串": "SKU008", "广式小腊肠": "SKU008",
    "鸡肉小串": "SKU009", "嫩滑鸡肉串": "SKU009",
    "雪花鸡柳串": "SKU010",
    # SKU011 卤山川大片鸡肉串：老王确认 黄金大里脊/香嫩大里脊 属于此 SKU
    "卤山川大片鸡肉串": "SKU011",
    "黄金大里脊": "SKU011",
    "香嫩大里脊": "SKU011",
    # 羊肉串
    "纯手工羊肉串": "SKU012", "草原羊肉串": "SKU012", "手穿羊肉串": "SKU012",
    # 年糕
    "脆皮年糕": "SKU013", "脆皮水磨年糕": "SKU013", "水磨年糕": "SKU013",
    # 香肠
    "开花肠": "SKU014", "开花大香肠": "SKU014",
    # 豆腐串
    "包浆豆腐串": "SKU015", "脆皮包浆豆腐": "SKU015",
    # 鱿鱼类（深海整只鱿鱼 → SKU020 大鱿鱼，老王确认）
    "鱿鱼龙爪须": "SKU016",
    "深海整只鱿鱼": "SKU020",
    "大鱿鱼": "SKU020", "轰炸大鱿鱼": "SKU020", "整只大鱿鱼": "SKU020", "深海大鱿鱼": "SKU020",
    # 骨肉相连
    "骨肉相连": "SKU017",
    # SKU018 卤山川巴掌大鸡排：老王确认 厚切拉丝鸡排/鸡排普拉丝 属于此 SKU
    "卤山川巴掌大鸡排": "SKU018",
    "厚切拉丝鸡排": "SKU018",
    "鸡排普拉丝": "SKU018",
    # 面筋
    "面筋串": "SKU019", "旋风炸面筋": "SKU019", "旋风大面筋": "SKU019",
    # 烧饼
    "椒盐烧饼": "SKU021", "酥皮千层烧饼": "SKU021", "酥皮千层饼": "SKU021",
    # 苕皮
    "苕皮包鱿鱼": "SKU022", "耙糯鱿鱼苕皮": "SKU022",
    "苕皮酸豆角": "SKU023", "耙糯酸豆角苕皮": "SKU023", "酸豆角苕皮": "SKU023",
    # 鸡腿
    "卤山川生炸鸡腿串": "SKU024", "生炸大鸡腿": "SKU024",
    # 凤爪
    "糯凤爪串": "SKU025", "王牌糯凤爪": "SKU025", "炸卤糯凤爪": "SKU025",
    # 牛肉
    "鲜切牛肉串": "SKU026", "鲜切嫩牛肉": "SKU026",
    "牛肚串": "SKU027", "厚切牛肚串": "SKU027", "牛上脑小串": "SKU027",
    # 鸡翅
    "奥尔良翅中": "SKU028", "奥尔良鸡翅中": "SKU028",
    # 豆扣结
    "香酥豆结扣": "SKU029", "香酥豆扣结": "SKU029",
    # 牛腩/牛肉筋
    "钢签黄牛肉筋": "SKU030", "钢签黄牛腩": "SKU030",
    # 散货（kg 类）
    "鸡叉骨": "SKU031", "肉多大鸡叉骨": "SKU031", "黄金鸡叉骨": "SKU031", "生炸鸡叉骨": "SKU031",
    "锅巴土豆": "SKU032", "脆皮小土豆": "SKU032", "云南锅巴土豆": "SKU032",
    "台湾无骨鸡柳": "SKU033",
    # 蔬菜
    "金针菇": "SKU057", "香韧金针菇": "SKU057",
    "洋白菜": "SKU058", "脆甜洋白菜": "SKU058", "洋白菜有点甜": "SKU058",
    "娃娃菜": "SKU059", "香甜娃娃菜": "SKU059",
    "虎皮尖椒": "SKU060", "虎皮青椒": "SKU060",
    "花菜": "SKU061", "清甜散花菜": "SKU061",
    # 雪花芝士年糕 —— 按老王：脆皮年糕 + 芝士（芝士不计库存）
    "雪花芝士年糕": "SKU013",
}

# 加工品/套餐 → 展开成多个 SKU 消耗（每份）
# 每条规则：关键词 → [(sku_id, base_qty), ...]
RECIPE_EXPAND = {
    # ---- 特例覆盖（优先于单 SKU + 规格解析）----
    # 炸卤糯凤爪（2只）→ 1 串糯凤爪串（老王确认 2只 = 1串）
    "炸卤糯凤爪": [("SKU025", 1)],

    # 单点组合商品（淘宝闪购）
    "厚切拉丝鸡排1串+脆皮水磨年糕1串+旋风大面筋1串": [
        ("SKU018", 1), ("SKU013", 1), ("SKU019", 1),
    ],
    "脆皮水磨年糕2串+旋风大面筋2串+香嫩大里脊1串": [
        ("SKU013", 2), ("SKU019", 2), ("SKU011", 1),
    ],

    # 烧饼夹里脊类：1 椒盐烧饼 + 2 串大片鸡肉串
    "烧饼夹里脊": [("SKU021", 1), ("SKU011", 2)],
    "酥皮烧饼夹里脊": [("SKU021", 1), ("SKU011", 2)],
    "烧饼夹里脊（双倍里脊）": [("SKU021", 1), ("SKU011", 4)],
    # 烧饼夹鸡柳：1 椒盐烧饼 + 50g 台湾无骨鸡柳
    "烧饼夹鸡柳": [("SKU021", 1), ("SKU033", 0.05)],
    # 炸串拌饼（3荤4素）：1 烧饼 + 2 串鱼豆腐 + 1 串骨肉相连 + 1 串大里脊
    "炸串拌饼（3荤4素）": [("SKU021", 1), ("SKU002", 2), ("SKU017", 1), ("SKU011", 1)],
    # 【专属特惠】炸串拌饼（鱼豆腐+里脊）：1 烧饼 + 2 串鱼豆腐 + 1 串大里脊
    "【专属特惠】炸串拌饼": [("SKU021", 1), ("SKU002", 2), ("SKU011", 1)],
    # 炸串标配四件套：1 大鸡排 + 1 年糕 + 1 面筋（饮料忽略）
    "炸串标配四件套": [("SKU018", 1), ("SKU013", 1), ("SKU019", 1)],
    # 酥皮汉堡单人餐：老王说鸡柳酥皮汉堡忽略，套餐父级也忽略
}

# 散货（kg 类）的规格专用消耗表（kg）
KG_SPEC = {
    # 鸡叉骨（老王 2026-05-12 更新：标准份 400g，生炸大桶 800g）
    ("SKU031", "肉多大鸡叉骨（6块）"): 0.40,
    ("SKU031", "肉多大鸡叉骨"): 0.40,   # 规格=6块/份
    ("SKU031", "黄金鸡叉骨"): 0.40,
    ("SKU031", "【满满超1斤】生炸鸡叉骨1大桶（生重超）"): 0.80,   # 老王 2026-05-12 指定 800g/份
    ("SKU031", "鸡叉骨"): 0.40,

    # 锅巴土豆（老王 2026-05-12 更新：1 串 = 30g，小土豆 15 个 = 180g）
    ("SKU032", "锅巴土豆串"): 0.03,        # 1 串 = 30g，再乘菜名里的"N串"
    ("SKU032", "脆皮小土豆（15个）"): 0.18,
    ("SKU032", "脆皮小土豆"): 0.18,        # 1人份/1盒 ≈ 15 个 ≈ 180g
    ("SKU032", "云南锅巴土豆1份（15块）"): 0.18,
    ("SKU032", "脆皮小土豆（2串）"): 0.06,  # 2 串 × 30g
    ("SKU032", "锅巴土豆"): 0.18,

    # 台湾无骨鸡柳：标准=150g，2份装=300g
    ("SKU033", "台湾无骨鸡柳（2份装）"): 0.30,
    ("SKU033", "台湾无骨鸡柳（份）"): 0.15,
    ("SKU033", "台湾无骨鸡柳"): 0.15,
}


# 忽略规则
IGNORE_KEYWORDS = [
    # 已下架的产品（老王确认）
    "牛上脑小串",
    "【口味】", "辣度", "含包装打包费", "打包费",
    "微微辣", "微辣", "中辣", "特辣", "不辣",
    "矿泉水", "可乐", "雪碧", "芬达", "杏皮茶",
    "套餐口味", "炸串口味",
    # 盲盒类（老王：不计库存）
    "盲盒", "随机福利", "福利购", "收藏购", "体验购",
    "1.99元购", "0.99元购", "0元加购", "1.99元福利购",
    "1.99元收藏购", "0.99元收藏购", "0.99元体验购",
    "拆0元", "【惊喜盲盒】",
    # 忽略的产品（老王决定）
    "南美白虾", "鸡柳酥皮汉堡", "里脊酥皮汉堡", "脆皮淀粉肠",
    "酥皮汉堡单人餐",   # 父级套餐，老王说鸡柳酥皮汉堡忽略即父级也忽略
    "新鲜蔬菜",          # 规格不明
    "土豆片",            # 规格=标准/份，量小（5 份），先忽略
    # 京东秒送特殊诱饵单品（和盲盒同类）
    "鸡排普拉丝（1串）",  # 京东秒送 95 份 --/份，老王答复归入 SKU018=鸡排，但独立诱饵单品，归盲盒一并忽略？
]

# ❗ 上面"鸡排普拉丝（1串）"老王明确说归入 SKU018，所以不应忽略。
# 修正做法：不要放进 IGNORE_KEYWORDS，而是让它走正常 NAME_TO_SKU 匹配。
IGNORE_KEYWORDS.remove("鸡排普拉丝（1串）")


# ============================================================
#  匹配工具
# ============================================================

def is_ignored(name: str):
    for kw in IGNORE_KEYWORDS:
        if kw in name:
            return f"ignore:{kw}"
    if name in ("不辣", "微辣", "中辣", "特辣", "微微辣", "标准"):
        return "ignore:口味选项"
    return None


def find_recipe(name: str):
    """加工品/套餐展开，返回 [(sku_id, base_qty per 份), ...] 或 None"""
    for kw, rec in RECIPE_EXPAND.items():
        if kw in name:
            return rec
    return None


def find_sku(name: str):
    """找到单 SKU，返回 (sku_id, 匹到的关键词) 或 None。优先长关键词。"""
    best = None
    for kw, sku_id in NAME_TO_SKU.items():
        if kw in name:
            if best is None or len(kw) > len(best[1]):
                best = (sku_id, kw)
    return best


def parse_qty_from_name(name: str) -> int | None:
    """从菜名里取"（N串）/（N只）/（N块）/（N个）"，返回 N。"""
    m = re.search(r"[（(](\d+)\s*(串|只|块|个|支|片)[）)]", name)
    if m:
        return int(m.group(1))
    return None


def parse_qty_from_spec(spec: str) -> tuple[float, str] | None:
    """从规格里取数字+单位，返回 (数字, 单位)。"""
    if not spec:
        return None
    m = re.match(r"(\d+(?:\.\d+)?)\s*(串|只|块|个|支|片|把|克|g|kg|人份|份)", spec)
    if m:
        return float(m.group(1)), m.group(2)
    return None


def base_qty_for_single_sku(sku_id: str, name: str, spec: str, unit: str) -> tuple[float, str]:
    """单 SKU 消耗量计算。返回 (base_qty_per_份, 备注)。"""
    sku = SKUS[sku_id]
    base_unit = sku["base_unit"]

    # === kg 类（散货）优先查专用表 ===
    if base_unit == "kg":
        # 先精确匹配菜名
        for (sid, kw), kg in KG_SPEC.items():
            if sid == sku_id and kw == name:
                return kg, f"散货规格表: {kw} = {kg}kg"
        # 包含匹配（最长优先）
        best = None
        for (sid, kw), kg in KG_SPEC.items():
            if sid != sku_id: continue
            if kw in name:
                if best is None or len(kw) > len(best[0]):
                    best = (kw, kg)
        if best:
            kw, kg = best
            # 特殊处理：锅巴土豆"N串"规则（1 串 = 30g）
            if sku_id == "SKU032" and "锅巴土豆串" in name:
                n = parse_qty_from_name(name) or 1
                return 0.03 * n, f"锅巴土豆 {n} 串 × 30g = {0.03*n}kg"
            return kg, f"散货规格表: {kw} = {kg}kg"
        # 规格含"N克/Ng"
        p = parse_qty_from_spec(spec)
        if p:
            n, u = p
            if u in ("克", "g"):
                return n / 1000, f"规格 {n}g"
            if u == "kg":
                return n, f"规格 {n}kg"
        # 默认 0.15kg/份
        return 0.15, "默认 1份=0.15kg"

    # === 串类 ===
    # 1) 菜名里明写"（N串）" 最优先
    n = parse_qty_from_name(name)
    if n is not None:
        return n, f"菜名含（{n}串）"

    # 2) 规格列里是"N串/N只/N把"
    p = parse_qty_from_spec(spec)
    if p:
        n, u = p
        if u in ("串", "只", "片", "个", "支"):
            return n, f"规格={spec}"

    # 特例：规格="串" 或 unit="串"（没带数字）→ 按 1 串
    if spec == "串":
        return 1, f"规格='串' → 默认 1 串"

    # 特例：SKU008 广式小腊肠 → 默认 5 串/份（规格=标准/--/空时；菜名或规格写了 10串会在前面步骤优先匹配）
    if sku_id == "SKU008" and spec in ("标准", "--", "1个", "", None) and not parse_qty_from_name(name):
        return 5, f"广式小腊肠默认 5 串/份"
        if u == "把":
            # 把 → 串 的系数查 sale_units
            try:
                sale = json.loads(sku["sale_units"])
                if "把" in sale:
                    return float(sale["把"]) * n, f"规格 {n}把 × {sale['把']}串/把"
            except Exception:
                pass
            return 0, f"规格有'把'但 sale_units 未定义"
        if u == "人份":
            return 0, f"spec='{spec}' 无法推断"

    # 3) 规格=标准/--/1串/空 → 默认 1 串
    if spec in ("标准", "--", "1串", "", None) or not spec:
        return 1, f"默认 1 串（spec={spec!r}）"

    # 4) unit=串
    if unit in ("串", "1串"):
        return 1, f"unit=串"
    if unit == "把":
        try:
            sale = json.loads(sku["sale_units"])
            if "把" in sale:
                return float(sale["把"]), f"unit=把 = {sale['把']}串"
        except Exception:
            pass

    return 0, f"无法解析 spec='{spec}' unit='{unit}'"


# ============================================================
#  主流程
# ============================================================

def main():
    rows = load_sales()
    mapped = []     # 最终销售映射（已展开套餐/加工品）
    ignored = []
    need_decision = []
    unmapped = []

    for r in rows:
        name = r["name"]

        # 1) 忽略
        ig = is_ignored(name)
        if ig:
            ignored.append({**r, "reason": ig})
            continue

        # 2) 加工品/套餐展开
        recipe = find_recipe(name)
        if recipe:
            for sku_id, per_unit_qty in recipe:
                mapped.append({
                    "platform": r["source"],
                    "item_name": r["name"],
                    "spec": r["spec"],
                    "unit": r["unit"],
                    "sku_id": sku_id,
                    "sku_name": SKUS[sku_id]["sku_name"],
                    "base_unit": SKUS[sku_id]["base_unit"],
                    "base_qty": per_unit_qty,
                    "qty": r["qty"],
                    "total_consumption": per_unit_qty * r["qty"],
                    "item_type": r["item_type"],
                    "note": f"加工品/套餐拆解: {name}",
                })
            continue

        # 3) 单 SKU 匹配
        hit = find_sku(name)
        if not hit:
            unmapped.append(r)
            continue
        sku_id, kw = hit
        base_qty, note = base_qty_for_single_sku(sku_id, name, r["spec"], r["unit"])
        if base_qty <= 0:
            need_decision.append({**r, "sku_id": sku_id, "matched_kw": kw, "note": note})
            continue
        mapped.append({
            "platform": r["source"],
            "item_name": r["name"],
            "spec": r["spec"],
            "unit": r["unit"],
            "sku_id": sku_id,
            "sku_name": SKUS[sku_id]["sku_name"],
            "base_unit": SKUS[sku_id]["base_unit"],
            "base_qty": base_qty,
            "qty": r["qty"],
            "total_consumption": base_qty * r["qty"],
            "item_type": r["item_type"],
            "note": f"匹配关键词='{kw}', {note}",
        })

    # 写出
    def write(path, rows, fields):
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            for r in rows:
                w.writerow({k: r.get(k, "") for k in fields})

    write(DATA / "sales_map_draft.csv", mapped,
          ["platform", "item_name", "spec", "unit", "sku_id", "sku_name", "base_unit",
           "base_qty", "qty", "total_consumption", "item_type", "note"])
    write(DATA / "ignored.csv", ignored,
          ["source", "item_type", "name", "spec", "unit", "qty", "reason"])
    write(DATA / "need_decision.csv", need_decision,
          ["source", "item_type", "name", "spec", "unit", "qty", "sku_id", "matched_kw", "note"])
    write(DATA / "unmapped.csv", unmapped,
          ["source", "item_type", "name", "spec", "unit", "qty"])

    print(f"✅ 映射成功:      {len(mapped):>4} 行 / {sum(r['qty'] for r in mapped):>7.0f} 份记录")
    print(f"🗑️  已忽略:        {len(ignored):>4} 行 / {sum(r['qty'] for r in ignored):>7.0f} 份记录")
    print(f"❓ 需决策:        {len(need_decision):>4} 行 / {sum(r['qty'] for r in need_decision):>7.0f} 份记录")
    print(f"❌ 未匹配:        {len(unmapped):>4} 行 / {sum(r['qty'] for r in unmapped):>7.0f} 份记录")


if __name__ == "__main__":
    main()
